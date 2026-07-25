#!/usr/bin/env python3
# Reads the wedding Excel and writes hafaka/data.json (plaintext, git-ignored).
# The plaintext JSON is later encrypted by encrypt.mjs before anything is committed.
import json, os, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "ספקים ולוז.xlsx")
OUT = os.path.join(HERE, "data.json")

# schedule short-name -> (role label, full supplier name from contacts)
SUPPLIER_MAP = {
    "ביני": ("עיצוב", "ביני בוקה"),
    "מישל": ("שיער ואיפור", "מישל תורגמן"),
    "מורן": ("צילום סטילס", "מורן מעיין"),
    "שי ונטלי": ("וידאו", "שי ונטלי"),
    "שקד": ("אולם", "בון בון"),
    "צליל": ("פולרויד", "Take & Shake"),
    "בן": ("דיג׳יי", "Final Drum"),
    "ריקי": ("זמרת", "ריקי בן ארי"),
    "אלירן": ("רב", "רבנות"),
    "טריאולי": ("להקה", "טריולי"),
    "טריולי": ("להקה", "טריולי"),
    "ארינה": ("פרחים", "Sinteza"),
}


def clean(v):
    if v is None:
        return ""
    if isinstance(v, str):
        v = v.strip()
        return "" if v in ("????", "-") else v
    if isinstance(v, (datetime.time,)):
        return v.strftime("%H:%M")
    if isinstance(v, (datetime.datetime,)):
        return v.strftime("%H:%M")
    return str(v).strip()


wb = openpyxl.load_workbook(XLSX, data_only=True)

# ---- Contacts ----
contacts = []
ws = wb["אנשי קשר"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not any(r[:4]):
        continue
    category, supplier, contact, phone = (clean(r[0]), clean(r[1]), clean(r[2]), clean(r[3]))
    if not (category or supplier or contact):
        continue
    contacts.append({
        "category": category,
        "supplier": supplier,
        "contact": contact,
        "phone": phone,  # "" means no number yet
    })

# ---- Schedule ----
schedule = []
seen_tokens = []
ws = wb["לוז"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
for r in rows:
    if not any(r[:6]):
        continue
    who, time, loc, act, sup_raw, notes = (
        clean(r[0]), clean(r[1]), clean(r[2]), clean(r[3]), clean(r[4]), clean(r[5]),
    )
    if not (act or loc or sup_raw):
        continue
    tokens = [t.strip() for t in sup_raw.split(",") if t.strip()] if sup_raw else []
    for t in tokens:
        if t not in seen_tokens:
            seen_tokens.append(t)
    schedule.append({
        "who": who,
        "time": time,          # "" -> render as "לפני האירוע"
        "location": loc,
        "activity": act,
        "suppliers": tokens,   # short-name tokens
        "notes": notes,
    })

# ---- Vendor filter chips (only vendors that appear in the schedule) ----
suppliers = []
for tok in seen_tokens:
    role, name = SUPPLIER_MAP.get(tok, ("", tok))
    suppliers.append({"key": tok, "role": role, "name": name})

data = {"contacts": contacts, "schedule": schedule, "suppliers": suppliers}

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=1)

print(f"wrote {OUT}: {len(contacts)} contacts, {len(schedule)} schedule rows, {len(suppliers)} vendor chips")
